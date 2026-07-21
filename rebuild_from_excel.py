# -*- coding: utf-8 -*-
"""
유니버스 엑셀 -> 대시보드(index.html) 재생성
---------------------------------------------
커버리지 종목을 추가/삭제/수정했을 때 이 스크립트만 돌리면
index.html 의  const DATA / const FIN  블록이 통째로 갱신됩니다.

사용법:
    pip install openpyxl
    python rebuild_from_excel.py
    python refresh_live.py      # 이어서 시세/리포트/일정도 갱신 권장

전제:
  - 엑셀 'Universe' 시트 레이아웃(열 순서)이 유지되어야 합니다.
  - 종목별 상세 시트는 Universe 순서대로 [종목],[종목_모델링] 쌍으로 존재.
"""
import openpyxl, json, re, sys, os

XLSX = r"C:\Users\user\Desktop\캐리인턴_유니버스_신주현_2 (2) (2) (2).xlsx"
HTML_PATH = "index.html"

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def val(x):
    if x is None or x == "" or isinstance(x, str) or hasattr(x, "isoformat"):
        return None
    try:
        f = float(x)
        return None if f != f else round(f, 3)
    except (TypeError, ValueError):
        return None


def cell(x):
    if x is None:
        return ""
    if isinstance(x, float):
        return round(x, 3)
    if hasattr(x, "isoformat"):
        return str(x)
    return x


def build_universe(wb):
    ws = wb["Universe"]
    rows = [[cell(c) for c in r] for r in ws.iter_rows(values_only=True)]

    records, picks = [], []
    for r in rows:
        # 데이터행 판별: [1]=순번(int), [3]=대섹터, [5]=종목명
        if isinstance(r[1], int) and r[3] and r[5] and r[25]:
            records.append({
                "pickLabel": r[0], "rank": r[1], "pick2": r[2],
                "sector": r[3], "sub": r[4], "name": r[5],
                "score": val(r[7]), "fairMktcap": val(r[8]), "upsideOwn": val(r[9]),
                "ret1w": val(r[10]), "ret1m": val(r[11]), "ret3m": val(r[12]),
                "mktcap": val(r[13]),
                "per12mf": val(r[14]), "per26": val(r[15]), "per27": val(r[16]),
                "rev26": val(r[17]), "rev27": val(r[18]),
                "op26": val(r[19]), "op27": val(r[20]),
                "eps26": val(r[21]), "eps27": val(r[22]),
                "epsg26": val(r[23]), "epsg27": val(r[24]),
                "price": val(r[25]), "target": val(r[26]), "upsideCons": val(r[27]),
            })
        # 섹터별 Top pick + 한줄 코멘트: [2]=순번, [5]=종목, [7]=코멘트
        elif isinstance(r[2], int) and r[5] and r[7]:
            picks.append({"rank": r[2], "sector": r[3], "sub": r[4],
                          "name": r[5], "thesis": r[7]})
    return records, picks


def build_financials(wb, names):
    """종목별 상세 시트에서 분기/연간 손익 추출"""
    QS = ["1Q24","2Q24","3Q24","4Q24","1Q25","2Q25","3Q25","4Q25",
          "1Q26","2Q26E","3Q26E","4Q26E","1Q27E","2Q27E","3Q27E","4Q27E"]
    sheet_by_name = {ws.title: ws for ws in wb.worksheets}
    fin, missing = {}, []

    for i, nm in enumerate(names):
        ws = sheet_by_name.get(nm)
        if ws is None:
            # Universe 순서 기준 위치로 fallback (INDEX/Universe/new_raw_data 다음부터 2칸씩)
            idx = 3 + 2 * i
            ws = wb.worksheets[idx] if idx < len(wb.worksheets) else None
        if ws is None:
            missing.append(nm); continue

        rows = list(ws.iter_rows(values_only=True))
        hdr = next((ri for ri, r in enumerate(rows)
                    if r and "2024A" in [str(c) if c is not None else "" for c in r]), None)
        if hdr is None:
            missing.append(nm); continue

        hc = [(str(c) if c is not None else "") for c in rows[hdr]]
        colmap = {c: ci for ci, c in enumerate(hc) if c}

        def grab(label):
            for r in rows[hdr + 1: hdr + 45]:
                if r and r[0] == label:
                    return {k: val(r[ci]) for k, ci in colmap.items() if ci < len(r)}
            return {}

        rev, op = grab("매출액"), grab("영업이익")
        fin[nm] = {
            "annualHist": {"2024A": [rev.get("2024A"), op.get("2024A")],
                           "2025A": [rev.get("2025A"), op.get("2025A")]},
            "quarters": [{"q": q, "rev": rev.get(q), "op": op.get(q)}
                         for q in QS if rev.get(q) is not None],
        }
    return fin, missing


def main():
    if not os.path.exists(XLSX):
        print("엑셀을 찾을 수 없습니다:", XLSX); sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    records, picks = build_universe(wb)
    names = [r["name"] for r in records]
    print(f"유니버스 {len(records)}종목 / 섹터 Top pick {len(picks)}건")

    raw_fin, missing = build_financials(wb, names)

    # 연간표 = 실적(엑셀 상세시트) + 추정(Universe 26E/27E)
    by_name = {r["name"]: r for r in records}
    fin = {}
    for nm, f in raw_fin.items():
        rc = by_name[nm]
        h = f["annualHist"]
        fin[nm] = {
            "annual": [
                {"y": "2024A", "rev": h["2024A"][0], "op": h["2024A"][1]},
                {"y": "2025A", "rev": h["2025A"][0], "op": h["2025A"][1]},
                {"y": "2026E", "rev": rc["rev26"], "op": rc["op26"]},
                {"y": "2027E", "rev": rc["rev27"], "op": rc["op27"]},
            ],
            "quarters": f["quarters"],
        }

    data = {"asOf": "", "analyst": "신주현", "records": records, "sectorPicks": picks}

    html = open(HTML_PATH, encoding="utf-8").read()
    html, n1 = re.subn(r"const DATA = \{.*?\};\n",
                       "const DATA = " + json.dumps(data, ensure_ascii=False) + ";\n",
                       html, count=1, flags=re.S)
    html, n2 = re.subn(r"const FIN = \{.*?\};\n",
                       "const FIN = " + json.dumps(fin, ensure_ascii=False) + ";\n",
                       html, count=1, flags=re.S)
    if not (n1 and n2):
        print(f"블록 교체 실패 (DATA={n1}, FIN={n2})"); sys.exit(1)
    open(HTML_PATH, "w", encoding="utf-8").write(html)

    print(f"[OK] index.html 갱신 - DATA {len(records)}종목 / FIN {len(fin)}종목")
    if missing:
        print("  상세 재무 없음:", missing)
    print("\n이어서 실행 권장:  python refresh_live.py")


if __name__ == "__main__":
    main()
