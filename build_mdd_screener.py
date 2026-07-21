# -*- coding: utf-8 -*-
"""
유니버스 낙폭(MDD) 스크리너 시트 생성기
--------------------------------------
캐리인턴_유니버스 워크북을 복사해 둔 파일에 '낙폭스크리너' 시트를 만듭니다.

  - 종목/섹터/분류(Top·2nd·Beta)/점수/견적 : Universe 시트에서 그대로 승계
  - 시총·PER·매출·영익·EPS·EPS증가율        : new_raw_data(QuantiWise) 캐시값 승계
  - 5일·20일·YTD 수익률, 52주 고점대비 낙폭,
    MDD(52주 / YTD), 52주 저점대비 반등        : 네이버 일봉 시세로 계산
  - 매력도 점수                              : 시트 상단 가중치 셀 기반 엑셀 수식(수정 가능)

사용법:
    python build_mdd_screener.py            # 기본 경로
    python build_mdd_screener.py <src.xlsx> <dst.xlsx>
"""
import json, re, sys, time, urllib.request, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gl
from openpyxl.formatting.rule import ColorScaleRule

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SRC = r"C:\Users\user\Desktop\캐리인턴_유니버스_신주현_2 (2) (2) (2).xlsx"
DST = r"C:\Users\user\Desktop\캐리인턴_유니버스_MDD_20260721.xlsx"
UA = {"User-Agent": "Mozilla/5.0"}
SHEET = "낙폭스크리너"

ROW_FIRST, ROW_LAST = 6, 34          # Universe 시트의 종목 행 범위


# ─────────────────────────────── 시세 수집 ───────────────────────────────
def fetch_daily(code, start, end):
    """네이버 일봉. [(yyyymmdd, open, high, low, close, vol), ...]"""
    url = ("https://api.finance.naver.com/siseJson.naver?"
           f"symbol={code}&requestType=1&startTime={start}&endTime={end}&timeframe=day")
    raw = urllib.request.urlopen(
        urllib.request.Request(url, headers=UA), timeout=20).read().decode("cp949", "replace")
    rows = []
    for m in re.finditer(r'\["(\d{8})",\s*([\d.]+),\s*([\d.]+),\s*([\d.]+),\s*([\d.]+),\s*([\d.]+)', raw):
        d, o, h, l, c, v = m.groups()
        rows.append((d, float(o), float(h), float(l), float(c), float(v)))
    return rows


def max_drawdown(closes):
    """peak → trough 최대 낙폭(%, 음수)."""
    peak, mdd = closes[0], 0.0
    for c in closes:
        peak = max(peak, c)
        mdd = min(mdd, c / peak - 1)
    return mdd * 100


def price_metrics(code, today):
    """5D·20D·YTD 수익률, 52주 고저, 현재 낙폭, MDD(52주/YTD), 저점대비 반등."""
    start52 = (today - datetime.timedelta(days=400)).strftime("%Y%m%d")
    bars = fetch_daily(code, start52, today.strftime("%Y%m%d"))
    if len(bars) < 30:
        return None

    closes = [b[4] for b in bars]
    highs = [b[2] for b in bars]
    lows = [b[3] for b in bars]
    last = closes[-1]

    # 정확히 52주 구간만 사용
    cut52 = (today - datetime.timedelta(days=365)).strftime("%Y%m%d")
    i52 = next((i for i, b in enumerate(bars) if b[0] >= cut52), 0)
    c52, h52, l52 = closes[i52:], highs[i52:], lows[i52:]

    ytd0 = f"{today.year}0101"
    iy = next((i for i, b in enumerate(bars) if b[0] >= ytd0), 0)
    cy = closes[iy:]

    hi52, lo52 = max(h52), min(l52)
    return {
        "price": last,
        "r5": (last / closes[-6] - 1) * 100 if len(closes) > 5 else None,
        "r20": (last / closes[-21] - 1) * 100 if len(closes) > 20 else None,
        "ytd": (last / cy[0] - 1) * 100 if len(cy) > 1 else None,
        "hi52": hi52, "lo52": lo52,
        "dd_now": (last / hi52 - 1) * 100,          # 52주 고점 대비 현재 낙폭
        "mdd52": max_drawdown(c52),
        "mddytd": max_drawdown(cy) if len(cy) > 1 else None,
        "off_low": (last / lo52 - 1) * 100,         # 52주 저점 대비 반등
        "bars": len(bars),
    }


def naver_extra(code):
    """PBR·배당수익률·외국인비중·컨센 목표가 (네이버 통합 API)."""
    def num(s):
        if s is None:
            return None
        s = str(s).replace(",", "").replace("%", "").replace("배", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    d = json.loads(urllib.request.urlopen(
        urllib.request.Request(f"https://m.stock.naver.com/api/stock/{code}/integration",
                               headers=UA), timeout=20).read().decode("utf-8"))
    ti = {t["code"]: t.get("value") for t in d.get("totalInfos", [])}
    return {"pbr": num(ti.get("pbr")), "div": num(ti.get("dividendYieldRatio")),
            "foreign": num(ti.get("foreignRate"))}


# ─────────────────────────── 워크북에서 기존 데이터 ───────────────────────────
def read_universe(src):
    """Universe + new_raw_data(캐시값)에서 종목별 기본/재무 데이터 수집."""
    wbf = openpyxl.load_workbook(src)                    # 수식(참조 추적용)
    wbv = openpyxl.load_workbook(src, data_only=True)    # 캐시값
    uf, uv = wbf["Universe"], wbv["Universe"]
    raw = wbf["new_raw_data"]

    # new_raw_data 9행(종목명) → 8행(코드) 매핑
    name2code = {}
    for c in range(2, raw.max_column + 1):
        nm, cd = raw.cell(9, c).value, raw.cell(8, c).value
        if nm and cd:
            name2code.setdefault(str(nm).strip(), str(cd).lstrip("A"))

    out, group = [], None
    for r in range(ROW_FIRST, ROW_LAST + 1):
        name = uv.cell(r, 6).value          # F: 종목
        if not name:
            continue
        if uv.cell(r, 3).value:             # C: Top/2nd/Beta (병합·비어있으면 위 값 승계)
            group = uv.cell(r, 3).value
        v = lambda col: uv.cell(r, col).value
        out.append({
            "no": v(2), "group": group, "sector": v(4), "sub": v(5), "name": str(name).strip(),
            "code": name2code.get(str(name).strip()),
            "score": v(8), "target_cap": v(9),          # H 점수, I 견적(십억)
            "mktcap": v(14),                            # N 시총(십억)
            "per_fwd": v(15), "per26": v(16), "per27": v(17),
            "sales26": v(18), "sales27": v(19),
            "op26": v(20), "op27": v(21),
            "eps26": v(22), "eps27": v(23),
            "epsg26": v(24), "epsg27": v(25),
            "px_raw": v(26), "cons_tp": v(27),          # Z 주가(쿼티와이즈), AA 컨센 목표가
        })
    return out, wbv


# ─────────────────────────────── 시트 작성 ───────────────────────────────
HDR = [
    ("분류", 9), ("대분류", 8), ("소분류", 9), ("종목", 13), ("코드", 8),
    ("시가총액\n(십억원)", 11), ("점수", 6), ("견적\n(십억원)", 10), ("견적\n상승여력(%)", 11),
    ("현재가\n(원)", 10), ("5일\n(%)", 8), ("20일\n(%)", 8), ("YTD\n(%)", 8),
    ("52주\n고점(원)", 11), ("고점대비\n낙폭(%)", 11), ("MDD\n52주(%)", 10), ("MDD\nYTD(%)", 10),
    ("52주저점\n대비(%)", 11),
    ("매출액\n26E", 10), ("매출액\n27E", 10), ("영업이익\n26E", 10), ("영업이익\n27E", 10),
    ("순이익(E)\n26E", 10), ("순이익(E)\n27E", 10),
    ("EPS\n26E", 9), ("EPS\n27E", 9), ("EPS증가율\n26E(%)", 11), ("EPS증가율\n27E(%)", 11),
    ("PER\n26E", 8), ("PER\n27E", 8), ("PER\nFwd12M", 9), ("PBR\n(TTM)", 8),
    ("컨센\n목표가", 10), ("컨센\n상승여력(%)", 11),
    ("매력도\n점수", 10), ("낙폭\n등급", 9),
]

THIN = Side(style="thin", color="FFBFBFBF")
MED = Side(style="medium", color="FF404040")
NAVY = "FF1F3864"
GREY = "FFEEEEEE"


def build(src=SRC, dst=DST):
    rows, wbv = read_universe(src)
    today = datetime.date.today()

    print(f"시세 수집 {len(rows)}종목 …")
    for i, s in enumerate(rows, 1):
        s.update({k: None for k in
                  ("price", "r5", "r20", "ytd", "hi52", "lo52", "dd_now",
                   "mdd52", "mddytd", "off_low", "pbr")})
        if not s["code"]:
            print(f"  [{i}/{len(rows)}] {s['name']}: 코드 없음 — 건너뜀")
            continue
        try:
            m = price_metrics(s["code"], today)
            if m:
                s.update(m)
            s["pbr"] = naver_extra(s["code"])["pbr"]
            print(f"  [{i}/{len(rows)}] {s['name']}({s['code']}) "
                  f"낙폭 {s['dd_now']:.1f}% / MDD52 {s['mdd52']:.1f}%")
        except Exception as e:
            print(f"  [{i}/{len(rows)}] {s['name']}({s['code']}) 실패: {str(e)[:60]}")
        time.sleep(0.2)

    wb = openpyxl.load_workbook(dst)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 2)          # INDEX, Universe 다음
    ws.sheet_properties.tabColor = "C00000"

    # ── 제목 & 가중치 박스 ──────────────────────────────────────────────
    ws["B2"] = f"소비재/유통/미용/음식료/엔터/게임/호텔 유니버스 — 낙폭과다 스크리너"
    ws["B2"].font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
    ws["B3"] = f"시세 기준일 {today:%Y-%m-%d} (네이버 일봉) · 재무 컨센서스 기준 new_raw_data({wbv['new_raw_data']['B1'].value})"
    ws["B3"].font = Font(name="맑은 고딕", size=9, color="FF808080")

    ws["B5"] = "매력도 가중치 (파란 셀 수정 가능, 합 100)"
    ws["B5"].font = Font(name="맑은 고딕", size=10, bold=True, color=NAVY)
    wlab = ["낙폭(고점대비)", "밸류(PER 26E)", "성장(EPS증가율 26E)", "견적 상승여력", "애널리스트 점수"]
    wval = [30, 20, 20, 20, 10]
    for j, (lb, vl) in enumerate(zip(wlab, wval)):
        c1, c2 = ws.cell(6, 2 + j), ws.cell(7, 2 + j)
        c1.value, c2.value = lb, vl
        c1.font = Font(name="맑은 고딕", size=9, color="FF404040")
        c1.alignment = Alignment(horizontal="center", wrap_text=True)
        c2.font = Font(name="맑은 고딕", size=11, bold=True, color="FF0000FF")
        c2.fill = PatternFill("solid", fgColor="FFDDEBF7")
        c2.alignment = Alignment(horizontal="center")
        c2.border = Border(THIN, THIN, THIN, THIN)
    ws["G7"] = "=SUM(B7:F7)"
    ws["G7"].font = Font(name="맑은 고딕", size=11, bold=True)
    ws["G6"] = "합계"
    ws["G6"].font = Font(name="맑은 고딕", size=9, color="FF404040")
    ws["G6"].alignment = Alignment(horizontal="center")

    # ── 헤더 ────────────────────────────────────────────────────────────
    HR = 10                                  # 헤더 행
    R0 = HR + 1                              # 첫 데이터 행
    for j, (t, w) in enumerate(HDR):
        c = ws.cell(HR, 2 + j)
        c.value = t
        c.font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(THIN, THIN, MED, MED)
        ws.column_dimensions[gl(2 + j)].width = w
    ws.row_dimensions[HR].height = 32

    # ── 데이터 ──────────────────────────────────────────────────────────
    rows.sort(key=lambda s: (s["dd_now"] if s["dd_now"] is not None else 0))   # 낙폭 큰 순
    n = len(rows)
    RN = R0 + n - 1

    def put(r, col, val, fmt=None, bold=False, color="FF000000", size=9):
        c = ws.cell(r, col)
        c.value = val
        c.font = Font(name="맑은 고딕", size=size, bold=bold, color=color)
        if fmt:
            c.number_format = fmt
        c.alignment = Alignment(horizontal="center" if col > 5 else "left", vertical="center")
        c.border = Border(THIN, THIN, THIN, THIN)
        return c

    F1, FP, FX = "#,##0.0", "0.0", "0.0"
    for k, s in enumerate(rows):
        r = R0 + k
        shift = 0 if s["price"] is None or not s["mktcap"] else 0
        put(r, 2, s["group"], size=9, bold=True,
            color={"Top pick": "FFC00000", "2nd pick": "FFBF8F00"}.get(s["group"], "FF404040"))
        put(r, 3, s["sector"])
        put(r, 4, s["sub"])
        put(r, 5, s["name"], bold=True)
        put(r, 6, s["code"])
        put(r, 7, s["mktcap"], "#,##0")
        put(r, 8, s["score"], "0.0")
        put(r, 9, s["target_cap"], "#,##0")
        put(r, 10, f"=IF(N({gl(9)}{r})=0,\"\",({gl(9)}{r}/{gl(7)}{r}-1)*100)", FP)
        put(r, 11, s["price"], "#,##0")
        put(r, 12, s["r5"], FP)
        put(r, 13, s["r20"], FP)
        put(r, 14, s["ytd"], FP)
        put(r, 15, s["hi52"], "#,##0")
        put(r, 16, s["dd_now"], FP, bold=True)
        put(r, 17, s["mdd52"], FP)
        put(r, 18, s["mddytd"], FP)
        put(r, 19, s["off_low"], FP)
        put(r, 20, s["sales26"], "#,##0")
        put(r, 21, s["sales27"], "#,##0")
        put(r, 22, s["op26"], "#,##0")
        put(r, 23, s["op27"], "#,##0")
        # 순이익(E) ≈ EPS × 주식수, 주식수 = 시총 / 현재가
        for cc, eps in ((24, "Z"), (25, "AA")):
            put(r, cc, f'=IF(OR({eps}{r}=0,L{r}=0,L{r}=""),"",{eps}{r}*H{r}*10^9/L{r}/10^9)', "#,##0")
        put(r, 26, s["eps26"], "#,##0")
        put(r, 27, s["eps27"], "#,##0")
        put(r, 28, s["epsg26"], FP)
        put(r, 29, s["epsg27"], FP)
        put(r, 30, s["per26"], "0.0\"x\"")
        put(r, 31, s["per27"], "0.0\"x\"")
        put(r, 32, s["per_fwd"], "0.0\"x\"")
        put(r, 33, s["pbr"], "0.00\"x\"")
        put(r, 34, s["cons_tp"] or None, "#,##0")
        put(r, 35, f'=IF(OR(N(AG{r})=0,N(L{r})=0),"",(AG{r}/L{r}-1)*100)', FP)

        # 매력도: 각 항목 백분위(0~1) × 가중치
        dd = f"PERCENTRANK.INC($Q${R0}:$Q${RN},Q{r},4)"          # 낙폭 클수록(값 작을수록) → 1-pct
        per = f"PERCENTRANK.INC($AE${R0}:$AE${RN},AE{r},4)"      # PER 낮을수록 좋음 → 1-pct
        gro = f"PERCENTRANK.INC($AC${R0}:$AC${RN},AC{r},4)"      # EPS 증가율 높을수록 좋음
        upd = f"PERCENTRANK.INC($J${R0}:$J${RN},J{r},4)"         # 견적 상승여력 높을수록 좋음
        put(r, 36,
            f'=IFERROR(ROUND('
            f'$B$7*(1-{dd})+'
            f'$C$7*IF(N(AE{r})<=0,0.5,1-{per})+'
            f'$D$7*{gro}+'
            f'$E$7*{upd}+'
            f'$F$7*(I{r}/10),1),"")',
            "0.0", bold=True, size=10)
        # 낙폭 등급
        put(r, 37, f'=IF(Q{r}="","",IF(Q{r}<=-40,"심각",IF(Q{r}<=-25,"과다",IF(Q{r}<=-15,"조정","정상"))))')

    # ── 조건부 서식 ──────────────────────────────────────────────────────
    ws.conditional_formatting.add(
        f"Q{R0}:Q{RN}", ColorScaleRule(start_type="num", start_value=-60, start_color="FFF8696B",
                                       mid_type="num", mid_value=-25, mid_color="FFFFEB84",
                                       end_type="num", end_value=0, end_color="FF63BE7B"))
    ws.conditional_formatting.add(
        f"AJ{R0}:AJ{RN}", ColorScaleRule(start_type="min", start_color="FFFFFFFF",
                                         end_type="max", end_color="FF63BE7B"))
    for col in ("M", "N", "O"):
        ws.conditional_formatting.add(
            f"{col}{R0}:{col}{RN}",
            ColorScaleRule(start_type="min", start_color="FFF8696B",
                           mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                           end_type="max", end_color="FF63BE7B"))

    ws.freeze_panes = f"G{R0}"
    ws.auto_filter.ref = f"B{HR}:AK{RN}"
    ws.sheet_view.showGridLines = False

    # ── 범례 ────────────────────────────────────────────────────────────
    lr = RN + 2
    for i, t in enumerate([
        "· 고점대비 낙폭(%) = 현재가 / 52주 최고가 - 1  →  '낙폭과다' 1차 스크린 지표 (기본 정렬 기준)",
        "· MDD 52주 / YTD  = 해당 구간 종가 기준 peak→trough 최대 낙폭 (변동성·하방 리스크 성격)",
        "· 순이익(E)는 EPS × 추정주식수(시총/현재가)로 역산한 값 — 정밀 수치는 각 종목 모델링 시트 참조",
        "· PBR은 네이버 TTM 기준, 그 외 재무는 new_raw_data(QuantiWise) 컨센서스 캐시값",
        "· 매력도 점수 = 낙폭·밸류·성장·상승여력 백분위 가중합 + 애널리스트 점수 (상단 B7:F7 가중치 수정 가능)",
        "· 낙폭 등급: 심각 ≤ -40% / 과다 ≤ -25% / 조정 ≤ -15% / 그 외 정상",
    ]):
        c = ws.cell(lr + i, 2)
        c.value = t
        c.font = Font(name="맑은 고딕", size=9, color="FF808080")

    wb.save(dst)
    print(f"\n[OK] '{SHEET}' 시트 생성 완료 → {dst}")
    return dst


if __name__ == "__main__":
    a = sys.argv[1:]
    build(a[0] if a else SRC, a[1] if len(a) > 1 else DST)
